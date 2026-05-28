import os
import shutil
import tempfile
import uuid

from fastapi import APIRouter, HTTPException, File, UploadFile, Body, Form
from fastapi.responses import StreamingResponse
import csv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from packages.utils import Logger, supabase_client, storage_client, BUCKET_NAME
from packages.file_storage import StorageClient
from packages.pdf_service import fetch_and_upload_pdf

logger = Logger()
router = APIRouter(tags=["conferences"])
storage_service = StorageClient(client=storage_client, bucket_name=BUCKET_NAME)

@router.post("/conferences/{conf_id}/banners")
async def add_conference_banner(
    conf_id: int,
    file: UploadFile = File(...)
):
    temp_file_path = None
    try:
        if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
             raise HTTPException(status_code=400, detail="Only accept images extension (png, jpg, jpeg, webp)")

        file_ext = file.filename.split(".")[-1]
        unique_filename = f"{uuid.uuid4()}.{file_ext}"
        
        gcs_path = f"conferences/{conf_id}/assets/{unique_filename}"

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_file_path = os.path.join(temp_dir, unique_filename)
            with open(temp_file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            public_url = storage_service.upload_blob(temp_file_path, gcs_path)
            
            if not public_url:
                raise HTTPException(status_code=500, detail="Error uploading GCS Storage")

      
        res = supabase_client.table("conferences").select("banner_urls").eq("conf_id", conf_id).single().execute()
        
        current_banners = res.data.get("banner_urls")
        
        if current_banners is None:
            current_banners = []
        
        current_banners.append(public_url)

        update_res = supabase_client.table("conferences").update({
            "banner_urls": current_banners
        }).eq("conf_id", conf_id).execute()

        return {
            "status": "success",
            "message": "banner added",
            "url": public_url,
            "all_banners": current_banners
        }

    except Exception as e:
        logger.error(f"Add Banner Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/conferences/{conf_id}/banners")
async def remove_conference_banner(
    conf_id: int,
    url_to_remove: str = Body(..., embed=True, description="Full URL of image to delete")
):
    
    try:
        res = supabase_client.table("conferences").select("banner_urls").eq("conf_id", conf_id).single().execute()
        current_banners = res.data.get("banner_urls")

        if not current_banners or url_to_remove not in current_banners:
            raise HTTPException(status_code=404, detail="URL not found or banner not in this conference")

        deleted = storage_service.delete_file(url_to_remove)
        
        if not deleted:
            logger.warning(f"File {url_to_remove} not found in GCS")

        new_banners = [url for url in current_banners if url != url_to_remove]

        supabase_client.table("conferences").update({
            "banner_urls": new_banners
        }).eq("conf_id", conf_id).execute()

        return {
            "status": "success",
            "message": "Banner removed sucessfully",
            "remaining_banners": new_banners
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Remove Banner Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/conferences/{conf_id}/import-history")
async def get_import_history(conf_id: int):
    try:
        res = (
            supabase_client.table("paper_imports")
            .select("*, person:profiles!person_in_charge(full_name)")
            .eq("conf_id", conf_id)
            .order("import_date", desc=True)
            .execute()
        )
        return {
            "status": "success",
            "data": res.data
        }
    except Exception as e:
        logger.error(f"Get Import History Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/conferences/{conf_id}/import-logs")
async def get_import_logs(conf_id: int):
    try:
        res = (
            supabase_client.table("paper_import_logs")
            .select("*, person:profiles!person_in_charge(full_name)")
            .eq("conf_id", conf_id)
            .order("created_at", desc=True)
            .limit(50)
            .execute()
        )
        return {
            "status": "success",
            "data": res.data
        }
    except Exception as e:
        logger.error(f"Get Import Logs Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/conferences/{conf_id}/import-template")
async def download_import_template(conf_id: int):
    """Generate and download a formatted Excel template for paper import."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Papers Import"

        # -- Styles --
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7")
        )
        example_font = Font(name="Calibri", size=10, italic=True, color="808080")

        # -- Headers --
        headers = [
            ("title", True, "Paper Title (Required)\n\nEnter the full title of the paper.\nThis field is required."),
            ("abstract", False, "Abstract (Optional)\n\nEnter a brief summary of the paper.\nCan be left empty."),
            ("primary_author_email", True, "Primary Author Email (Required)\n\nEmail of the main author.\nMust be a registered user in the system."),
            ("co_author_emails", False, "Co-author Emails (Optional)\n\nEmails of co-authors, separated by semicolons (;).\nExample: user1@mail.com;user2@mail.com\nAll emails must belong to registered users."),
            ("external_pdf_url", False, "External PDF URL (Optional)\n\nPaste a direct, publicly accessible download link to the PDF file.\nOnly links from Google Drive are supported.\nSystem will auto-download and store it.\nLeave empty to skip PDF upload for this row."),
        ]

        for col_idx, (header_name, is_required, comment_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = required_fill if is_required else header_fill
            cell.alignment = header_align
            cell.border = thin_border
            cell.comment = Comment(comment_text, "System")
            cell.comment.width = 300
            cell.comment.height = 120

        # -- Column widths --
        ws.column_dimensions["A"].width = 35  # title
        ws.column_dimensions["B"].width = 55  # abstract
        ws.column_dimensions["C"].width = 30  # primary_author_email
        ws.column_dimensions["D"].width = 40  # co_author_emails
        ws.column_dimensions["E"].width = 50  # external_pdf_url

        # -- Example row --
        examples = [
            "Example: AI in Healthcare",
            "Example: This paper discusses the role of AI...",
            "example@email.com",
            "coauthor1@email.com;coauthor2@email.com",
            "https://example.com/paper.pdf",
        ]
        for col_idx, val in enumerate(examples, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = example_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # -- Empty rows with borders for user to fill --
        for row_idx in range(3, 22):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # -- Freeze header row --
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

        # -- Instructions sheet --
        ws_info = wb.create_sheet(title="Instructions")
        ws_info.column_dimensions["A"].width = 80
        instructions = [
            ("📋 Paper Import Template - Instructions", Font(name="Calibri", bold=True, size=14, color="2F5496")),
            ("", None),
            ("1. Fill in the 'Papers Import' sheet with your paper data.", Font(name="Calibri", size=11)),
            ("2. Required fields: title, primary_author_email (columns highlighted in darker blue).", Font(name="Calibri", size=11)),
            ("3. Optional fields: abstract, co_author_emails.", Font(name="Calibri", size=11)),
            ("4. All email addresses must belong to registered users in the system.", Font(name="Calibri", size=11, color="CC0000")),
            ("5. For multiple co-authors, separate emails with semicolons (;).", Font(name="Calibri", size=11)),
            ("6. Row 2 contains examples — delete or overwrite it before uploading.", Font(name="Calibri", size=11)),
            ("7. Status will automatically be set to 'ACCEPTED' for all imported papers.", Font(name="Calibri", size=11)),
            ("8. For external_pdf_url, ONLY Google Drive share links are supported.", Font(name="Calibri", size=11, color="CC0000")),
            ("", None),
            ("⚠️ Important: Do NOT change the column headers in the first row.", Font(name="Calibri", bold=True, size=11, color="CC0000")),
        ]
        for row_idx, (text, font) in enumerate(instructions, 1):
            cell = ws_info.cell(row=row_idx, column=1, value=text)
            if font:
                cell.font = font
            cell.alignment = Alignment(wrap_text=True)

        # -- Save to buffer --
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=paper_import_template_conf_{conf_id}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Generate Template Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/conferences/{conf_id}/import-papers")
async def import_papers_csv(
    conf_id: int,
    uploader_id: int = Form(...),
    file: UploadFile = File(...)
):
    filename_lower = file.filename.lower()
    if not filename_lower.endswith(('.csv', '.xlsx')):
        raise HTTPException(status_code=400, detail="Only accept .csv or .xlsx files")
    
    try:
        content = await file.read()
        
        # Parse rows based on file type
        if filename_lower.endswith('.xlsx'):
            # Parse Excel file
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            all_rows_raw = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if len(all_rows_raw) < 2:
                raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows.")
            
            # First row = headers
            raw_headers = [str(h).strip().lower() if h else "" for h in all_rows_raw[0]]
            required_columns = ["title", "abstract", "primary_author_email", "co_author_emails"]
            if not all(col in raw_headers for col in required_columns):
                raise HTTPException(
                    status_code=400,
                    detail=f"Excel must contain the following headers: {', '.join(required_columns)}"
                )
            
            rows = []
            for data_row in all_rows_raw[1:]:
                row_dict = {}
                for col_idx, header in enumerate(raw_headers):
                    val = data_row[col_idx] if col_idx < len(data_row) else None
                    row_dict[header] = str(val).strip() if val is not None else ""
                # Skip completely empty rows
                if any(row_dict.get(col) for col in required_columns[:1]):  # at least title
                    rows.append(row_dict)
        else:
            # Parse CSV file
            try:
                text_content = content.decode('utf-8-sig')
            except UnicodeDecodeError:
                text_content = content.decode('latin-1')
            
            csv_reader = csv.DictReader(io.StringIO(text_content))
            required_columns = ["title", "abstract", "primary_author_email", "co_author_emails"]
            if not csv_reader.fieldnames or not all(col in csv_reader.fieldnames for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"CSV must contain the following headers: {', '.join(required_columns)}"
                )
            rows = list(csv_reader)
        
        # Phase 1: Validation
        # Collect all emails to fetch
        all_emails = set()
        for i, row in enumerate(rows):
            if row.get("primary_author_email"):
                all_emails.add(row["primary_author_email"].strip())
            
            co_authors_str = row.get("co_author_emails", "")
            if co_authors_str:
                co_emails = [e.strip() for e in co_authors_str.split(";") if e.strip()]
                for ce in co_emails:
                    all_emails.add(ce)
                    
        # Fetch existing users
        profiles_res = supabase_client.table("profiles").select("user_id, email").in_("email", list(all_emails)).execute()
        existing_users = {p["email"]: p["user_id"] for p in profiles_res.data}
        
        # Fetch existing papers in this conference for duplicate check
        existing_papers_res = supabase_client.table("papers").select("title").eq("submitted_conf", conf_id).execute()
        existing_titles = {p["title"].strip().lower() for p in existing_papers_res.data} if existing_papers_res.data else set()
        
        # Check for missing users + duplicates
        all_errors = []
        valid_rows = []
        for i, row in enumerate(rows):
            row_num = i + 2 # +1 for 0-index, +1 for header
            row_has_error = False
            
            # Check duplicate title
            title = row.get("title", "").strip()
            if title and title.lower() in existing_titles:
                all_errors.append(f"Row {row_num}: Paper '{title}' already exists in this conference.")
                row_has_error = True
            
            primary_email = row.get("primary_author_email", "").strip()
            
            if not primary_email:
                all_errors.append(f"Row {row_num}: Missing primary_author_email")
                row_has_error = True
            elif primary_email not in existing_users:
                all_errors.append(f"Row {row_num}: Primary author email '{primary_email}' does not exist in the system.")
                row_has_error = True
            
            co_authors_str = row.get("co_author_emails", "")
            if co_authors_str:
                co_emails = [e.strip() for e in co_authors_str.split(";") if e.strip()]
                
                # Check overlap between primary author and co-authors
                if primary_email in co_emails:
                    all_errors.append(f"Row {row_num}: Primary author '{primary_email}' cannot be listed as a co-author.")
                    row_has_error = True
                
                # Check for duplicate co-authors in the same row
                if len(co_emails) != len(set(co_emails)):
                    all_errors.append(f"Row {row_num}: Duplicate co-author emails found in the co-authors list.")
                    row_has_error = True
                
                for ce in co_emails:
                    if ce not in existing_users:
                        all_errors.append(f"Row {row_num}: Co-author email '{ce}' does not exist in the system.")
                        row_has_error = True
            
            if not row_has_error:
                row["_original_row_num"] = row_num
                valid_rows.append(row)

        # Phase 2: Insertion
        imported_count = 0

        for row in valid_rows:
            row_num = row["_original_row_num"]
            title = row.get("title", "").strip()
            abstract = row.get("abstract", "").strip()
            primary_email = row.get("primary_author_email", "").strip()
            primary_id = existing_users[primary_email]

            # Map optional table fields
            status = row.get("status", "").strip() or "ACCEPTED"
            final_decision_date = row.get("final_decision_date", "").strip() or None
            created_at_val = row.get("created_at", "").strip() or None

            insert_data = {
                "title": title,
                "abstract": abstract,
                "primary_author_id": primary_id,
                "submitted_conf": conf_id,
                "status": status,
            }
            if final_decision_date:
                insert_data["final_decision_date"] = final_decision_date
            if created_at_val:
                insert_data["created_at"] = created_at_val

            # Insert paper
            paper_res = supabase_client.table("papers").insert(insert_data).execute()

            if not paper_res.data:
                continue

            paper_id = paper_res.data[0]["paper_id"]

            # Insert co-authors
            co_authors_str = row.get("co_author_emails", "")
            if co_authors_str:
                co_emails = [e.strip() for e in co_authors_str.split(";") if e.strip()]
                coauthor_records = [
                    {"paper_id": paper_id, "user_id": existing_users[ce], "author_order": idx + 1}
                    for idx, ce in enumerate(co_emails)
                ]
                if coauthor_records:
                    supabase_client.table("paper_coauthors").insert(coauthor_records).execute()

            # --- PDF auto-download from external link ---
            external_pdf_url = row.get("external_pdf_url", "").strip()
            if external_pdf_url:
                try:
                    if "1drv.ms" in external_pdf_url or "onedrive" in external_pdf_url.lower():
                        raise ValueError("OneDrive links are not supported. Please use Google Drive.")
                        
                    # Create a placeholder version record first to get version_id
                    version_res = supabase_client.table("paper_versions").insert({
                        "paper_id": paper_id,
                        "version_number": 1,
                        "upload_by": uploader_id,
                        "is_final": False,
                        "format_ok": False,
                        "display": True,
                        "file_path": "pending_upload",
                    }).execute()

                    version_id = version_res.data[0]["version_id"]

                    gcs_url = await fetch_and_upload_pdf(
                        url=external_pdf_url,
                        paper_id=str(paper_id),
                        version_id=str(version_id),
                    )

                    supabase_client.table("paper_versions").update({
                        "file_path": gcs_url
                    }).eq("version_id", version_id).execute()

                    logger.info(f"PDF linked for paper_id={paper_id}, version_id={version_id}: {gcs_url}")

                except Exception as pdf_err:
                    err_msg = f"Row {row_num} (Paper '{title}'): PDF download failed from '{external_pdf_url}' — {pdf_err}"
                    logger.error(err_msg)
                    all_errors.append(err_msg)
                    
                    # ROLLBACK: delete the inserted paper
                    supabase_client.table("papers").delete().eq("paper_id", paper_id).execute()
                    
                    continue

            imported_count += 1
            
        # Log successful imports if any
        if imported_count > 0:
            supabase_client.table("paper_imports").insert({
                "conf_id": conf_id,
                "num_papers": imported_count,
                "file_name": file.filename,
                "person_in_charge": uploader_id
            }).execute()

        # If there are any errors, we log them
        if all_errors:
            error_text = " | ".join(all_errors)
            supabase_client.table("paper_import_logs").insert({
                "conf_id": conf_id,
                "status": "WARNING" if imported_count > 0 else "ERROR",
                "file_name": file.filename,
                "num_papers": imported_count,
                "error_details": error_text,
                "person_in_charge": uploader_id
            }).execute()
            
            if imported_count == 0:
                raise HTTPException(status_code=400, detail=error_text)
        else:
            # Full success logging
            supabase_client.table("paper_import_logs").insert({
                "conf_id": conf_id,
                "status": "SUCCESS",
                "file_name": file.filename,
                "num_papers": imported_count,
                "error_details": None,
                "person_in_charge": uploader_id
            }).execute()

        return {
            "status": "success",
            "message": f"Successfully imported {imported_count} papers",
            "count": imported_count,
            "errors": all_errors
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Import Papers Error: {e}")
        # Log unexpected errors
        try:
            supabase_client.table("paper_import_logs").insert({
                "conf_id": conf_id,
                "status": "ERROR",
                "file_name": file.filename,
                "num_papers": 0,
                "error_details": str(e),
                "person_in_charge": uploader_id
            }).execute()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=str(e))

